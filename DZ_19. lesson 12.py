import openpyxl
import csv

with open('name_age.csv', 'r', encoding='utf-8') as file:
    name_age = csv.reader(file, delimiter=',')

    wb = openpyxl.Workbook()
    wb.create_sheet(title="DZ_19", index=0)
    sheet = wb['DZ_19']
    rows = list(name_age)

    for i, row in enumerate(rows, start=1):
        sheet.cell(row=1, column=i+1, value='Person'+str(i))
        for j, value in enumerate(row):
            sheet.cell(row=j + 2, column=i, value=value)
    cell = sheet.cell(row=1, column=i+1)
    cell.value = None

    for i in range(1, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=i, column=j)
            if cell.value == 'Возраст':
                sheet.delete_rows(cell.row)

    wb.save('DZ_19.xlsx')
    wb.close()

